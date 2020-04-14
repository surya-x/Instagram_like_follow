from config import *
import openpyxl
import sys
import shutil
import os

# To give the all the details in "parameter.xlsx"
def retrieve_file_parameter():
    logging.info("retrieve file parameter called")
    try:
        workbook = openpyxl.load_workbook(r"assets\parameters.xlsx")
        sheet = workbook.worksheets[0]

        path = sheet["A2"].value + r"\insta_search_found.xlsx"
        follow_limit = sheet["B2"].value
        time_limit = sheet["C2"].value

        return path, follow_limit, time_limit
    except Exception as e:
        print("\nError : loading data from parameters.xlsx")
        print(e)
        logging.error(e)
        os.system("PAUSE")
        sys.exit()


# To give the no. of rows in insta_search_found.xlsx
def get_nrows(path):
    logging.info("get_nrows called")
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.worksheets[0]
        nrows = sheet.max_row
        return nrows
    except Exception as e:
        print("Error : loading 'insta_search_found.xlsx' ")
        print(e)
        logging.error(e)
        os.system("PAUSE")
        sys.exit()


# Make sure that atleast 1 row (with information) is available to use bot
def check_rows(nrows):
    logging.info("check_rows called")
    if nrows < 2:
        print("No username given in 'insta_search_found.xlsx' ")


# Make sure atleast 1 rows without "OK" status is available to use bot
# True for any Blank column left to execute
# False for NO blank column left ( All OK )
def check_ok_status(path):
    logging.info("check ok status called")
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.worksheets[0]
        no_of_rows = sheet.max_row

        ok_count = 0

        for i in range(2, no_of_rows+1):
            if sheet.cell(row=i, column=2).value is not None:
                ok_count = ok_count + 1

        if ok_count == (no_of_rows - 1):
            print("Status of all username is 'OK' or Filled ")
            rename_file(path)
            sys.exit()                                                          # quits the code
        else:
            print("check_ok_status passed")                                     # LOGS
    except Exception as e:
        print("Error : loading 'insta_search_found.xlsx' ")
        print(e)
        logging.error(e)
        os.system("PAUSE")
        sys.exit()


# TO read the username and status from path("insta_search_found.xlsx") and return a list
def read_usernames(path):
    logging.info("read usernames called")
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.worksheets[0]

        username_list = []
        ok_list = []
        no_of_rows = sheet.max_row

        for i in range(2, no_of_rows + 1):
            username_list.append(sheet.cell(row=i, column=1).value)
            ok_list.append(sheet.cell(row=i, column=2).value)

        return username_list, ok_list
    except Exception as e:
        print("Error : loading 'insta_search_found.xlsx' ")
        print(e)
        logging.error(e)
        os.system("PAUSE")
        sys.exit()                                                          # QUIT HERE


def write_ok_status(path, row):
    logging.info("write ok status called")
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.worksheets[0]

        sheet.cell(row=row, column=2).value = "OK"
        workbook.save(path)
    except Exception as e:
        print("Error : Writing OK in insta_seach_done.xlsx")
        logging.error(e)

def wrong_status(path, row):
    logging.info("wrong status called")
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.worksheets[0]

        sheet.cell(row=row, column=2).value = "WRONG"
        workbook.save(path)
    except Exception as e:
        print("Error : Writing in insta_seach_done.xlsx")
        logging.error(e)


# Will rename the file "insta_search_found.xlsx" into "insta_search_found.xlsx"
def rename_file(path):
    logging.info("rename file called")
    try:
        workbook = openpyxl.load_workbook(r"assets\parameters.xlsx")
        sheet = workbook.worksheets[0]

        newpath = sheet["A2"].value + r"\insta_search_found_done.xlsx"
        shutil.move(path, newpath)
        print("\n...DONE...")
    except Exception as e:
        print("Error : renaming 'insta_search_found.xlsx' ")
        print(e)
        logging.error(e)
